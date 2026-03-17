# core/core_engine.py
import os
import pandas as pd
import openpyxl
from datetime import datetime, timedelta

class CoreEngine:
    """核心引擎 - 处理表格筛选和区分逻辑"""
    
    def __init__(self, app):
        self.app = app
        
        # 排除列表
        self.exclude_list = ['炬', '荷花', '金兔']
        
        # 账号列名可能性
        self.account_cols = ['会员账号', '会员名', '账号', '指定代理', '会员帐号', '指定', '代理']
        
        # 过滤关键词
        self.filter_keywords = ['统计', '总计', '合计', '汇总', 'Total']
        
        # ===== 正确的写入标签映射 =====
        self.write_tags = {
            # D表写入 - 对应D表处理结果
            "D": {
                "D表30钱包": {"mode": "full", "reward": 3, "deposit": 30, "count": 1},
                "D表100钱包": {"mode": "full", "reward": 6, "deposit": 100, "count": 1},
                "D表1000钱包": {"mode": "full", "reward": 38, "deposit": 1000, "count": 1},
                "D表站内信": {"mode": "simple"}
            },
            # C表写入 - 对应C表处理结果
            "C": {
                "C表未使用APP": {"mode": "full", "reward": 3.88, "deposit": 0, "count": 0},
                "C表站内信": {"mode": "simple"}
            },
            # A表写入 - 对应A表处理结果
            "A": {
                "A表充值未提": {"mode": "A_special"},
                "A表站内信": {"mode": "simple"}
            },
            # B表写入 - 对应B表处理结果
            "B": {
                "B表50任务": {"mode": "dispatch", "custom_data": {"fixed": {2: 8, 3: 50, 4: 1}}},
                "B表300任务": {"mode": "dispatch", "custom_data": {"fixed": {2: 20, 3: 300, 4: 1}}},
                "B表站内信": {"mode": "simple"}
            },
            # H表写入 - 对应H表处理结果
            "H": {
                "H表任务": {"mode": "dispatch", "custom_data": {"fixed": {2: 8, 3: 50, 4: 1}}},
                "H表站内信": {"mode": "simple"}
            },
            # 3表写入 - 对应3表处理结果
            "3": {
                "3神秘彩金": {"mode": "dispatch", "custom_data": {"col_map": {2: "金币"}, "fixed": {3: 0, 4: 0}}},
                "3幸运彩金": {"mode": "dispatch", "custom_data": {"col_map": {2: "金币"}, "fixed": {3: 0, 4: 0}}},
                "3神秘站内信": {"mode": "simple"},
                "3幸运站内信": {"mode": "simple"}
            }
        }
    
    def _calculate_bonus(self, amount):
        """阶梯彩金计算逻辑"""
        try:
            amount = float(amount)
        except:
            return 0
            
        if amount >= 10000: return 188.88
        if amount >= 5000: return 118.88
        if amount >= 3000: return 88.88
        if amount >= 1000: return 58.88
        if amount >= 700: return 38.88
        if amount >= 500: return 18.88
        if amount >= 100: return 8.88
        if amount >= 0: return 1.88
        return 0
    
    def _clean_dataframe(self, df, source_name="未知"):
        """清洗DataFrame，统一列名并过滤统计行"""
        if df is None:
            return None
            
        try:
            # 找出账号列
            current_cols = {str(c).strip(): c for c in df.columns}
            found_col = None
            for col in self.account_cols:
                if col in current_cols:
                    found_col = current_cols[col]
                    break
            
            if not found_col:
                found_col = df.columns[0]
                self.app.add_log("执行", "核心引擎", f"[{source_name}] 未匹配到标准表头，使用第1列作为账号")
            
            # 重命名列
            df = df.rename(columns={found_col: '会员账号'})
            
            # 如果有第二列，命名为金币
            if "金币" not in df.columns and len(df.columns) > 1:
                df = df.rename(columns={df.columns[1]: '金币'})
            
            # 过滤统计行
            mask = df['会员账号'].astype(str).str.contains('|'.join(self.filter_keywords), na=False)
            return df[~mask].copy()
            
        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"清洗 [{source_name}] 失败: {str(e)}")
            return None
    
    def _get_account_list(self, df, account_col='会员账号'):
        """从DataFrame获取账号列表"""
        if df is None or df.empty:
            return []
        return df[account_col].drop_duplicates().tolist()
    
    def _write_batch(self, prefix, account_list=None, df=None):
        """批量写入表格"""
        if prefix not in self.write_tags:
            return
            
        for tag, config in self.write_tags[prefix].items():
            # 如果是A表特殊模式且有DataFrame
            if prefix == "A" and df is not None and config["mode"] == "A_special":
                self.smart_excel_writer(tag, df, **config)
            # 如果有账号列表
            elif account_list:
                self.smart_excel_writer(tag, account_list, **config)
            # 如果有DataFrame且不是A_special模式
            elif df is not None and config["mode"] != "A_special":
                self.smart_excel_writer(tag, df, **config)
    
    # ========== 公开方法 ==========
    
    def smart_excel_writer(self, tag, account_list, mode="simple", reward=0, deposit=0, count=0, custom_data=None):
        """全兼容写入器：将数据写入对应的Excel文件"""
        success_count = 0
        
        try:
            if not hasattr(self.app, 'M1_data_storage'):
                self.app.add_log("警告", "核心引擎", "数据存储未初始化")
                return

            file_path = self.app.M1_data_storage.get(tag)
            if not file_path or not os.path.exists(file_path):
                self.app.add_log("警告", "核心引擎", f"标签 [{tag}] 未导入，已跳过写入")
                return

            wb = openpyxl.load_workbook(file_path)
            ws = wb.active 
            
            # 清空旧数据
            if ws.max_row >= 2:
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
                    for cell in row:
                        cell.value = None
            
            if account_list is None or (hasattr(account_list, '__len__') and len(account_list) == 0):
                wb.save(file_path)
                self.app.add_log("执行", "核心引擎", f"📝 [{tag}] 数据为空，清空完成")
                return

            # DataFrame类型写入
            if hasattr(account_list, 'itertuples'):
                for i, row_data in enumerate(account_list.itertuples(), start=2):
                    acc = str(getattr(row_data, '会员账号', ''))
                    if any(kw in acc for kw in self.filter_keywords):
                        continue
                    
                    ws.cell(row=i, column=1, value=acc)
                    
                    if mode == "A_special":
                        bonus = self._calculate_bonus(getattr(row_data, '充值金额', 0))
                        ws.cell(row=i, column=2, value=bonus)
                        ws.cell(row=i, column=3, value=0)
                        ws.cell(row=i, column=4, value=0)
                    
                    elif mode == "dispatch" and custom_data:
                        col_map = custom_data.get("col_map", {})
                        fixed_vals = custom_data.get("fixed", {})
                        
                        for col_idx, src_col in col_map.items():
                            ws.cell(row=i, column=col_idx, value=getattr(row_data, src_col, None))
                        
                        for col_idx, f_val in fixed_vals.items():
                            ws.cell(row=i, column=col_idx, value=f_val)
                    
                    success_count += 1
            
            # 列表类型写入
            else:
                for i, acc in enumerate(account_list, start=2):
                    acc_str = str(acc)
                    if any(kw in acc_str for kw in self.filter_keywords):
                        continue
                    
                    ws.cell(row=i, column=1, value=acc_str)
                    
                    if mode == "full":
                        ws.cell(row=i, column=2, value=reward)
                        ws.cell(row=i, column=3, value=deposit)
                        ws.cell(row=i, column=4, value=count)
                    
                    success_count += 1

            wb.save(file_path)
            self.app.add_log("成功", "核心引擎", f"📝 [{tag}] 写入成功 (共计 {success_count} 条)")

        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"❌ 写入 {tag} 异常: {str(e)}")

    def export_simple_list_to_desktop(self, prefix, data):
        """导出简单列表到桌面"""
        if data is None or (isinstance(data, list) and not data):
            return

        try:
            yesterday = datetime.now() - timedelta(days=1)
            date_str = f"{yesterday.month}月{yesterday.day}日"
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            file_name = f"{date_str}{prefix}.xlsx"
            save_path = os.path.join(desktop, file_name)

            if prefix == "A" and hasattr(data, 'itertuples'):
                report_list = []
                for row in data.itertuples():
                    bonus = self._calculate_bonus(getattr(row, '充值金额', 0))
                    report_list.append({
                        "会员账号": getattr(row, '会员账号', ''),
                        "彩金": bonus,
                        "打码量": bonus,
                        "业务类型": "其他优惠",
                        "备注": "幸运彩金",
                        "后台备注": "幸运彩金2"
                    })
                df_to_save = pd.DataFrame(report_list)
            else:
                if hasattr(data, 'columns') and '会员账号' in data.columns:
                    acc_list = data['会员账号'].tolist()
                else:
                    acc_list = data if isinstance(data, list) else []
                df_to_save = pd.DataFrame({"会员账号": acc_list})

            df_to_save.to_excel(save_path, index=False)
            self.app.add_log("成功", "核心引擎", f"💾 已导出 {prefix} 表至桌面: {file_name}")

        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"❌ {prefix} 表导出失败: {str(e)}")

    def append_to_source_file(self, tag, data):
        """追加数据到源文件"""
        if not hasattr(self.app, 'M1_data_storage'):
            self.app.add_log("失败", "核心引擎", "错误：路径字典未初始化")
            return

        file_path = self.app.M1_data_storage.get(tag)
        if not file_path or not os.path.exists(file_path):
            self.app.add_log("警告", "核心引擎", f"[{tag}] 尚未上传文件，跳过追加")
            return

        try:
            # 提取账号列表
            if hasattr(data, 'columns') and '会员账号' in data.columns:
                acc_list = data['会员账号'].tolist()
            else:
                acc_list = data if isinstance(data, list) else []
            
            if not acc_list:
                return

            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # 找到最后一行
            last_row = ws.max_row
            while last_row > 0:
                if ws.cell(row=last_row, column=1).value not in [None, ""]:
                    break
                last_row -= 1
            
            target_row = last_row + 1
            for acc in acc_list:
                ws.cell(row=target_row, column=1, value=str(acc))
                target_row += 1
            
            wb.save(file_path)
            self.app.add_log("成功", "核心引擎", f"♻️ 追加 {len(acc_list)} 个账号至 [{tag}]")
            
        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"❌ [{tag}] 追加失败: {str(e)}")
    
    # ========== 核心业务逻辑 ==========
    
    def process_d_table(self, df_fs, df_rc, threshold=0):
        """处理D表逻辑：首存未充"""
        if df_fs is None or df_rc is None:
            return False, []
            
        try:
            self.app.add_log("执行", "核心引擎", "执行 D表 (首存未充) 筛选...")
            
            if '充值渠道' not in df_rc.columns or '会员账号' not in df_rc.columns:
                self.app.add_log("失败", "核心引擎", "D表缺失必要列名")
                return False, []
            
            # 过滤充值渠道
            pattern = '|'.join(self.exclude_list)
            rc_filtered = df_rc[~df_rc['充值渠道'].str.contains(pattern, na=False)]
            rc_accs = set(rc_filtered['会员账号'].unique())
            
            # 找出首存但未充值的账号
            d_list = df_fs[~df_fs['会员账号'].isin(rc_accs)]['会员账号'].drop_duplicates().tolist()
            
            if d_list:
                self._write_batch("D", d_list)
                self.export_simple_list_to_desktop("D", d_list)
                self.app.add_log("成功", "核心引擎", f"D表处理完毕：剩余 {len(d_list)} 账号")
                return True, d_list
            else:
                self.app.add_log("执行", "核心引擎", "D表筛选结果为空")
                return False, []
                
        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"D表逻辑崩溃: {str(e)}")
            return False, []
    
    def process_c_table(self, df_fs, df_login):
        """处理C表逻辑：未用APP"""
        if df_fs is None or df_login is None:
            return False, []
            
        try:
            self.app.add_log("执行", "核心引擎", "执行 C表 (未用APP) 筛选...")
            
            if '设备端' not in df_login.columns or '会员账号' not in df_login.columns:
                self.app.add_log("失败", "核心引擎", "C表缺失必要列名")
                return False, []
            
            # 找出使用APP的用户
            app_users = set(df_login[df_login['设备端'].str.contains('APP', case=False, na=False)]['会员账号'].unique())
            
            # 找出首存但未用APP的账号
            c_list = df_fs[~df_fs['会员账号'].isin(app_users)]['会员账号'].drop_duplicates().tolist()
            
            if c_list:
                self._write_batch("C", c_list)
                self.export_simple_list_to_desktop("C", c_list)
                self.app.add_log("成功", "核心引擎", f"C表处理完毕：剩余 {len(c_list)} 账号")
                return True, c_list
            else:
                self.app.add_log("执行", "核心引擎", "C表筛选结果为空")
                return False, []
                
        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"C表逻辑崩溃: {str(e)}")
            return False, []
    
    def process_a_table(self, df_cz_not_ti, df_total_not_ti, percent):
        """处理A表逻辑：充提差额（亏损比例筛选）"""
        if df_cz_not_ti is None or df_total_not_ti is None:
            return False, []
            
        try:
            self.app.add_log("执行", "核心引擎", f"执行 A表 (负值比例筛选) 比例: {percent}...")
            
            required_cols = ['会员账号', '会员输赢', '充值金额']
            if not all(col in df_cz_not_ti.columns for col in required_cols):
                self.app.add_log("失败", "核心引擎", "A表缺失必要列名")
                return False, []
            
            if '会员账号' not in df_total_not_ti.columns:
                self.app.add_log("失败", "核心引擎", "总表未提缺失会员账号列")
                return False, []
            
            # 排除已在总表中的账号
            total_accs = set(df_total_not_ti['会员账号'].unique())
            df_filtered = df_cz_not_ti[~df_cz_not_ti['会员账号'].isin(total_accs)].copy()
            
            if df_filtered.empty:
                self.app.add_log("执行", "核心引擎", "A表账号均已在总表中，跳过")
                return False, []
            
            # 数据清洗
            df_filtered['会员输赢'] = pd.to_numeric(df_filtered['会员输赢'], errors='coerce').fillna(0)
            df_filtered['充值金额'] = pd.to_numeric(df_filtered['充值金额'], errors='coerce').fillna(0)
            
            # 解析比例
            try:
                raw_val = str(percent).replace('%', '').strip()
                threshold = abs(float(raw_val) / 100)
            except:
                threshold = 0.2
                self.app.add_log("警告", "核心引擎", f"比例解析失败，使用默认值 {threshold*100}%")
            
            # 计算亏损比例
            df_filtered['abs_ratio'] = df_filtered['会员输赢'].abs() / df_filtered['充值金额'].replace(0, float('inf'))
            
            # 筛选亏损且比例达标的账号
            temp_df = df_filtered[
                (df_filtered['会员输赢'] < 0) & 
                (df_filtered['abs_ratio'] >= threshold)
            ].copy()
            
            if temp_df.empty:
                self.app.add_log("执行", "核心引擎", f"A表无符合比例≥{threshold*100}%的账号")
                return False, []
            
            # 过滤统计行
            mask = temp_df['会员账号'].astype(str).str.contains('|'.join(self.filter_keywords), na=False)
            final_df = temp_df[~mask].copy()
            
            if final_df.empty:
                self.app.add_log("执行", "核心引擎", "A表过滤统计行后为空")
                return False, []
            
            # 提取账号列表
            a_list = final_df['会员账号'].drop_duplicates().tolist()
            
            if a_list:
                self._write_batch("A", a_list, final_df)
                self.export_simple_list_to_desktop("A", final_df)
                self.append_to_source_file("总表未提", a_list)
                self.app.add_log("成功", "核心引擎", f"A表处理完毕：筛选出亏损≥{threshold*100}% 的账号 {len(a_list)} 个")
                return True, a_list
            else:
                self.app.add_log("执行", "核心引擎", "A表账号列表为空")
                return False, []
                
        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"A表逻辑异常: {str(e)}")
            return False, []
    
    def process_b_table(self, df_b):
        """处理B表逻辑：原始数据B"""
        df_clean = self._clean_dataframe(df_b, "原始数据B")
        if df_clean is None:
            return False
        
        try:
            self.app.add_log("执行", "核心引擎", "处理原始数据B...")
            self._write_batch("B", None, df_clean)
            self.app.add_log("成功", "核心引擎", "原始数据B处理完成 ✅")
            return True
        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"原始数据B处理崩溃: {str(e)}")
            return False
    
    def process_h_table(self, df_h):
        """处理H表逻辑：原始数据H"""
        df_clean = self._clean_dataframe(df_h, "原始数据H")
        if df_clean is None:
            return False
        
        try:
            self.app.add_log("执行", "核心引擎", "处理原始数据H...")
            self._write_batch("H", None, df_clean)
            self.app.add_log("成功", "核心引擎", "原始数据H处理完成 ✅")
            return True
        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"原始数据H处理崩溃: {str(e)}")
            return False
    
    def process_3_table(self, df_3):
        """处理3表逻辑：原始数据3 (神秘/幸运分发)"""
        df_clean = self._clean_dataframe(df_3, "原始数据3")
        if df_clean is None:
            return False
        
        try:
            self.app.add_log("执行", "核心引擎", "处理原始数据3...")
            
            if '备注' not in df_clean.columns:
                self.app.add_log("失败", "核心引擎", "原始数据3缺失 '备注' 列，无法分流")
                return False
            
            executed = False
            
            # 神秘分支
            df_mystic = df_clean[df_clean['备注'].astype(str).str.contains("神秘彩金")].copy()
            if not df_mystic.empty:
                self._write_batch("3", None, df_mystic)
                self.app.add_log("成功", "核心引擎", f"神秘彩金分支处理完成 ({len(df_mystic)} 账号)")
                executed = True
            
            # 幸运分支
            df_lucky = df_clean[df_clean['备注'].astype(str).str.contains("幸运彩金")].copy()
            if not df_lucky.empty:
                self._write_batch("3", None, df_lucky)
                self.app.add_log("成功", "核心引擎", f"幸运彩金分支处理完成 ({len(df_lucky)} 账号)")
                executed = True
            
            if executed:
                self.app.add_log("成功", "核心引擎", "原始数据3处理完成 ✅")
                return True
            else:
                self.app.add_log("执行", "核心引擎", "原始数据3无匹配数据")
                return False
                
        except Exception as e:
            self.app.add_log("系统", "核心引擎", f"原始数据3处理崩溃: {str(e)}")
            return False
    
    def execute_alpha(self, percent):
        """执行表格筛选逻辑 (ALPHA)"""
        vault = self.app.data_vault
        
        # 获取数据
        df_fs = vault.get("昨日首存")
        df_rc = vault.get("昨日充值")
        df_login = vault.get("昨日登录")
        df_cz_not_ti = vault.get("充值未提")
        df_total_not_ti = vault.get("总表未提")
        
        executed = False
        
        # 执行D表
        success, _ = self.process_d_table(df_fs, df_rc)
        executed = executed or success
        
        # 执行C表
        success, _ = self.process_c_table(df_fs, df_login)
        executed = executed or success
        
        # 执行A表
        success, _ = self.process_a_table(df_cz_not_ti, df_total_not_ti, percent)
        executed = executed or success
        
        if executed:
            self.app.add_log("成功", "核心引擎", "表格筛选任务处理完毕 🏁")
        else:
            self.app.add_log("失败", "核心引擎", "未执行任何有效任务")
    
    def execute_beta(self):
        """执行表格区分逻辑 (BETA)"""
        vault = self.app.data_vault
        
        # 获取数据
        df_b = vault.get("原始数据B")
        df_h = vault.get("原始数据H")
        df_3 = vault.get("原始数据3")
        
        executed = False
        
        # 执行B表
        if self.process_b_table(df_b):
            executed = True
        
        # 执行H表
        if self.process_h_table(df_h):
            executed = True
        
        # 执行3表
        if self.process_3_table(df_3):
            executed = True
        
        if executed:
            self.app.add_log("成功", "核心引擎", "表格区分任务处理完毕 🏁")
        else:
            self.app.add_log("失败", "核心引擎", "未执行任何有效任务")
